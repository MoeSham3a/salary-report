"""
Excel Report Generator — Creates styled .xlsx reports matching the original layout.
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from salary_engine import calculate_employee_quarterly, calculate_employee_yearly, calculate_summary


# Styling constants
HEADER_FONT = Font(name='Tajawal', size=16, bold=True)
SUBHEADER_FONT = Font(name='Tajawal', size=11, bold=True)
LABEL_FONT = Font(name='Tajawal', size=10, bold=True)
DATA_FONT = Font(name='Tajawal', size=10)
TOTAL_FONT = Font(name='Tajawal', size=10, bold=True, color='FFFFFF')

HEADER_FILL = PatternFill('solid', fgColor='1E293B')
LABEL_FILL = PatternFill('solid', fgColor='334155')
TOTAL_FILL = PatternFill('solid', fgColor='4F46E5')
SUMMARY_FILL = PatternFill('solid', fgColor='F1F5F9')

RTL_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True,
                       readingOrder=2)  # RTL
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', readingOrder=2)
NUM_ALIGN = Alignment(horizontal='right', vertical='center', readingOrder=2)

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)

# Column definitions for quarterly report (right-to-left order in Excel)
QUARTERLY_COLUMNS = [
    ('بدل نقل', 'transport', 14),
    ('الضريبة', 'tax', 14),
    ('المبلغ الخاضع', 'taxable', 16),
    ('التنزيل العائلي', 'family_deduction', 16),
    ('الفرق المستحق', 'amount_due', 14),
    ('مجموع الاشتراكات', 'total_contributions', 18),
    ('نهاية الخدمة 8.5%', 'end_of_service', 16),
    ('تعويضات عائلية 6%', 'family_6', 16),
    ('المرض و الامومة 11%', 'sickness_11', 18),
    ('التعويضات العائلية المدفوعة', 'family_paid', 20),
    ('مجموع الرواتب', 'total_salary', 16),
    ('الراتب الشهري', 'monthly_salary', 14),
    ('عدد الأولاد', 'num_children', 10),
    ('متزوج', 'is_married', 8),
    ('اجنبي', 'is_foreign', 8),
    ('تاريخ بدء العمل', 'start_date', 14),
    ('اسم الأجير', 'name', 20),
    ('رقم', 'row_number', 6),
]

YEARLY_COLUMNS = [
    ('بدل نقل', 'transport', 14),
    ('الضريبة', 'tax', 14),
    ('المبلغ الخاضع', 'taxable', 16),
    ('التنزيل العائلي', 'family_deduction', 16),
    ('الفرق المستحق', 'amount_due', 14),
    ('الإشتراكات 25.5%', 'contributions_25', 16),
    ('التعويضات العائلية المدفوعة', 'family_paid', 20),
    ('مجموع الرواتب', 'total_salary', 16),
    ('الراتب الشهري', 'monthly_salary', 14),
    ('عدد الأولاد', 'num_children', 10),
    ('متزوج', 'is_married', 8),
    ('اجنبي', 'is_foreign', 8),
    ('تاريخ بدء العمل', 'start_date', 14),
    ('اسم الأجير', 'name', 20),
    ('رقم', 'row_number', 6),
]

SUMMARY_ROWS = [
    'مجموع الرواتب',
    'المنافع النقدية والعينية',
    'مجموع المبالغ المدفوعة',
    'تعويضات نقل وانتقال',
    'التعويضات العائلية',
    'مجموع الرواتب_net',
    'التنزيل العائلي',
    'المبلغ الخاضع',
    'الضريبة المتوجبة',
]

SUMMARY_LABELS = [
    'مجموع الرواتب',
    'المنافع النقدية والعينية',
    'مجموع المبالغ المدفوعة',
    'تعويضات نقل وانتقال',
    'التعويضات العائلية',
    'مجموع الرواتب',
    'التنزيل العائلي',
    'المبلغ الخاضع',
    'الضريبة المتوجبة',
]


def generate_report(employees, company, report_type='quarter'):
    """
    Generate Excel report.
    report_type: 'quarter' or 'year'
    Returns BytesIO with .xlsx data.
    """
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True  # RTL layout

    columns = QUARTERLY_COLUMNS if report_type == 'quarter' else YEARLY_COLUMNS
    calc_fn = calculate_employee_quarterly if report_type == 'quarter' else calculate_employee_yearly
    num_cols = len(columns)

    # Set column widths
    for i, (_, _, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Row 1: Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value='جدول الرواتب و الأجور')
    title_cell.font = HEADER_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center', reading_order=2)
    title_cell.fill = HEADER_FILL
    title_cell.font = Font(name='Tajawal', size=16, bold=True, color='FFFFFF')
    ws.row_dimensions[1].height = 35

    # --- Row 2: Company info ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    period_label = f"عن فصل {company.get('quarter', 1)}" if report_type == 'quarter' else f"عن سنة {company.get('year', 2026)}"
    info_text = f"إسم المؤسسة: {company.get('name', '')}  |  الرقم المالي: {company.get('financial_number', '')}  |  رقم الضمان: {company.get('social_security_number', '')}  |  {period_label}"
    info_cell = ws.cell(row=2, column=1, value=info_text)
    info_cell.font = Font(name='Tajawal', size=10, color='64748B')
    info_cell.alignment = Alignment(horizontal='center', vertical='center', reading_order=2)
    ws.row_dimensions[2].height = 25

    # --- Row 3: Empty spacer ---
    ws.row_dimensions[3].height = 8

    # --- Row 4: Column headers ---
    for i, (label, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=4, column=i, value=label)
        cell.font = LABEL_FONT
        cell.font = Font(name='Tajawal', size=9, bold=True, color='FFFFFF')
        cell.fill = LABEL_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 30

    # --- Employee data rows ---
    data_start_row = 5
    for emp_idx, emp in enumerate(employees):
        row_num = data_start_row + emp_idx
        calc = calc_fn(emp)

        for col_idx, (_, key, _) in enumerate(columns, 1):
            if key in calc:
                value = calc[key]
            elif key in emp:
                value = emp[key]
            else:
                value = ''

            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = NUM_ALIGN if isinstance(value, (int, float)) else RTL_ALIGN
            cell.border = THIN_BORDER

            if isinstance(value, (int, float)) and key not in ('row_number', 'num_children'):
                cell.number_format = '#,##0'

        # Alternate row shading
        if emp_idx % 2 == 1:
            for col_idx in range(1, num_cols + 1):
                ws.cell(row=row_num, column=col_idx).fill = PatternFill('solid', fgColor='F8FAFC')

    # --- TOTAL row ---
    total_row = data_start_row + len(employees)
    # "TOTAL" label in the name column (second-to-last column)
    name_col_idx = num_cols - 1  # اسم الأجير column
    ws.cell(row=total_row, column=name_col_idx, value='المجموع').font = TOTAL_FONT
    ws.cell(row=total_row, column=name_col_idx).fill = TOTAL_FILL
    ws.cell(row=total_row, column=name_col_idx).alignment = CENTER_ALIGN
    ws.cell(row=total_row, column=name_col_idx).border = THIN_BORDER

    # Sum numeric columns
    all_calcs = [calc_fn(emp) for emp in employees]
    for col_idx, (_, key, _) in enumerate(columns, 1):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THIN_BORDER
        cell.alignment = NUM_ALIGN

        if key in ('name', 'start_date', 'is_foreign', 'is_married', 'row_number'):
            continue

        total_val = sum(c.get(key, 0) or 0 for c in all_calcs)
        if key == 'monthly_salary':
            total_val = sum(emp.get('monthly_salary', 0) or 0 for emp in employees)
        elif key == 'num_children':
            total_val = sum(emp.get('num_children', 0) or 0 for emp in employees)

        cell.value = total_val
        cell.number_format = '#,##0'

    # --- Summary section ---
    summary = calculate_summary(employees, report_type)
    summary_start = total_row + 2

    for i, (key, label) in enumerate(zip(SUMMARY_ROWS, SUMMARY_LABELS)):
        row = summary_start + i

        # Label in columns spanning 2 cols
        label_cell = ws.cell(row=row, column=num_cols - 3, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = RTL_ALIGN
        label_cell.fill = SUMMARY_FILL
        label_cell.border = THIN_BORDER

        ws.merge_cells(start_row=row, start_column=num_cols - 3, 
                       end_row=row, end_column=num_cols - 2)

        # Value
        val_cell = ws.cell(row=row, column=num_cols - 4, value=summary.get(key, 0))
        val_cell.font = Font(name='Tajawal', size=10, bold=True)
        val_cell.alignment = NUM_ALIGN
        val_cell.border = THIN_BORDER
        val_cell.number_format = '#,##0'

        if label in ('مجموع الرواتب', 'المبلغ الخاضع', 'الضريبة المتوجبة'):
            val_cell.font = Font(name='Tajawal', size=10, bold=True, color='4F46E5')

    # Write to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
